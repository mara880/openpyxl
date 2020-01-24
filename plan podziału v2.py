from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
arkusz = Workbook()
skoroszyt = arkusz.active
skoroszyt.title = "Plan podziału"
def naglowek():
    skoroszyt.merge_cells('A1:A3')
    skoroszyt.cell(row=1, column=1).value = "Nazwa Wierzyciela"
    skoroszyt.merge_cells('B1:B3')
    skoroszyt.cell(row=1, column=2).value = "nr TW"
    skoroszyt.merge_cells('C1:C3')
    skoroszyt.cell(row=1, column=3).value = "Należność główna"
    skoroszyt.merge_cells('D1:D3')
    skoroszyt.cell(row=1, column=4).value = "Odsetki"
    skoroszyt.merge_cells('E1:E3')
    skoroszyt.cell(row=1, column=5).value = "Koszty upomnienia"
    skoroszyt.merge_cells('F1:F3')
    skoroszyt.cell(row=1, column=6).value = "Koszty egzekucyjne"
    skoroszyt.merge_cells('G1:G3')
    skoroszyt.cell(row=1, column=7).value = "Udział procentowy (należność główna + odsetki"

def kategoria_1(row, numer_kategorii):
    skoroszyt.merge_cells(start_row=row, end_row=row, start_column=1, end_column=7)
    skoroszyt.cell(row=row, column=1).value = "Wierzyciel kategorii " + numer_kategorii

def wprowadzanie_danych(wiersz):
    nazwa_wierzyciela = str(input("Podaj nazwę wierzyciela:"))
    skoroszyt.cell(row = wiersz, column = 1).value = nazwa_wierzyciela
    numer_tw = str(input("Podaj numer TW:"))
    skoroszyt.cell(row = wiersz, column = 2).value = numer_tw
    nal_glowna = float(input("Podaj należność główną:"))
    skoroszyt.cell(row = wiersz, column = 3).value = nal_glowna
    odsetki = float(input("Podaj kwotę odsetek na dzień wpływu pieniędzy:"))
    skoroszyt.cell(row = wiersz, column = 4).value = odsetki
    koszty_upo = float(input("Podaj kwotę kosztów upomnienia:"))
    skoroszyt.cell(row = wiersz, column = 5).value = koszty_upo
    koszty_egz = float(input("Podaj kwotę kosztów egzekucyjnych:"))
    skoroszyt.cell(row = wiersz, column = 6).value = koszty_egz
        #zakoncz = input("Czy chcesz zakończyc dodawanie (T/N)?")
    wiersz = wiersz + 1

print('''W pierwszej kolejności wskaż liczbę wierzycieli
            następnie zacznij podawać wierzycieli według kategorii
            tj. od kategorii 1 do kategorii 7
            jeśli chcesz zakońćzyć wpisz T''')


liczba_kat_1 = int(input("Ilosc kat 1:"))
liczba_kat_2 = int(input("Ilosc kat 2:"))
liczba_kat_3 = int(input("Ilosc kat 3:"))
liczba_kat_4 = int(input("Ilosc kat 4:"))
liczba_kat_5 = int(input("Ilosc kat 5:"))
liczba_kat_6 = int(input("Ilosc kat 6:"))
liczba_wierzycieli = liczba_kat_1 + liczba_kat_2 + liczba_kat_3 + liczba_kat_4 + liczba_kat_5 + liczba_kat_6

naglowek()

wiersz_startowy = 4

licznik_pomocniczy = 0

if liczba_kat_1 > 0:
    kategoria_1(wiersz_startowy, str(1))                #tytul kategorii I

while licznik_pomocniczy < liczba_kat_1:
    print("Podajesz wierzycieli kategorii 1")
    wiersz_startowy += 1
    wprowadzanie_danych(wiersz_startowy)
    licznik_pomocniczy += 1
                                                            #starter kategorii II
if liczba_kat_2 > 0:
    if liczba_kat_1 == 0:
        liczba_kat_1 += 1
    kategoria_1(wiersz_startowy+liczba_kat_1, str(2))       #tytul kategorii II
    licznik_pomocniczy = 0
                                                            #wprowadzanie danych kategorii II
while licznik_pomocniczy < liczba_kat_2:
    print("Podajesz wierzycieli kategorii 2")
    wiersz_startowy = wiersz_startowy+liczba_kat_1+1
    wprowadzanie_danych(wiersz_startowy)
    licznik_pomocniczy += 1
                                                            #starter kategorii III
if liczba_kat_3 > 0:
    if liczba_kat_2 == 0:
        liczba_kat_2 =+ 1
    kategoria_1(wiersz_startowy+liczba_kat_2, str(3))       #tytul kategorii III

licznik_pomocniczy = 0
                                                            #wprowadzanie danych kategorii III
while licznik_pomocniczy < liczba_kat_3:
    print("Podajesz wierzycieli kategorii 3")
    wiersz_startowy = wiersz_startowy+liczba_kat_2+1
    wprowadzanie_danych(wiersz_startowy)
    licznik_pomocniczy += 1
                                                            #starter kategorii IV
if liczba_kat_4 > 0:
    if liczba_kat_3 == 0:
        liczba_kat_3 += 1
    kategoria_1(wiersz_startowy+liczba_kat_3, str(4))       #tytul kategorii IV

licznik_pomocniczy = 0
                                                            #wprowadzanie danych kategorii IV
while licznik_pomocniczy < liczba_kat_4:
    print("Podajesz wierzycieli kategorii 4")
    wiersz_startowy = wiersz_startowy+liczba_kat_3+1
    wprowadzanie_danych(wiersz_startowy)
    licznik_pomocniczy +=1
                                                            #starter kategorii V
if liczba_kat_5 > 0:
    if liczba_kat_4 == 0:
        liczba_kat_4 =+ 1
    kategoria_1(wiersz_startowy+liczba_kat_4, str(5))      #tytul kategorii V

licznik_pomocniczy = 0
                                                            #wprowadzanie danych kategorii V
while licznik_pomocniczy < liczba_kat_5:
    print("Podajesz wierzycieli kategorii 5")
    wiersz_startowy = wiersz_startowy+liczba_kat_4+1
    wprowadzanie_danych(wiersz_startowy)
    licznik_pomocniczy +=1
                                            #starter kategorii VI uruchamia się jeśli jest jakiś wierzyciel tej kategorii
                                            #jesli kategoria poprzednia(V) byla zerowa to dodaje jeden wiersz zeby nie nakłądaly się
                                            #tytuly.
                                            #starter zawsze sprawdza poprzednia kategorie czy posiada jakiegos wierzyciela
if liczba_kat_6 > 0:
    if liczba_kat_5 == 0:
        liczba_kat_5 += 1
    kategoria_1(wiersz_startowy+liczba_kat_5,str(6))        #tytul kategorii VI

licznik_pomocniczy = 0

while licznik_pomocniczy < liczba_kat_6:
    print("Podajesz wierzycieli kategorii 6")
    wiersz_startowy = wiersz_startowy+liczba_kat_5+1
    wprowadzanie_danych(wiersz_startowy)
    licznik_pomocniczy += 1

kwota_uzyskana = int(input("Podaj kwotę uzyskaną: "))
# skoroszyt.cell(row = 7, column =  4).value = "=SUMA(D4:D6)"

arkusz.save("plan v2.xlsx")

otwarcie = load_workbook(filename = "plan v2.xlsx")




